#!/usr/bin/env python3
"""
Import 🔵 High — Emerging Clinical accounts + contacts into Signal.

Join strategy:
  - Roll Numbers from "🔵 High — Emerging Clinical" sheet  (raw values, not formulas)
  - Contact + company data from "👥 Contacts (CRM Link)"   (raw data)
  → filter Contacts by matching Roll Numbers → 2,419 contact rows across ~444 companies

Signal tables populated:
  - organizations   (deduped by name — source data has no domain/website)
  - people          (deduped by linkedin_url; fallback name+org)
  - campaign_organizations + campaign_people  (only if --campaign is given)

Usage:
  python3 scripts/import-helyx-high.py <xlsx-path>                    # dry run
  python3 scripts/import-helyx-high.py <xlsx-path> --run              # actually write
  python3 scripts/import-helyx-high.py <xlsx-path> --run \\
      --campaign "Helyx High — Emerging Clinical"                      # + link campaign
"""

import sys
import os
import json
import re
import requests
import openpyxl
from pathlib import Path
from collections import defaultdict

# ── Config ────────────────────────────────────────────────────────────────────

XLSX_PATH    = sys.argv[1] if len(sys.argv) > 1 else None
DRY_RUN      = "--run" not in sys.argv
CAMPAIGN_ARG = None
if "--campaign" in sys.argv:
    idx = sys.argv.index("--campaign")
    CAMPAIGN_ARG = sys.argv[idx + 1] if idx + 1 < len(sys.argv) else None

SOURCE_LABEL = "helyx_crm_high"

# ── Load .env.local ───────────────────────────────────────────────────────────

def load_env(path=".env.local"):
    env = {}
    try:
        with open(path) as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, _, v = line.partition("=")
                # strip inline comments and quotes
                v = re.sub(r"\s+#.*$", "", v).strip().strip('"').strip("'")
                env[k.strip()] = v
    except FileNotFoundError:
        pass
    return env

env = load_env()
SUPABASE_URL      = env.get("NEXT_PUBLIC_SUPABASE_URL") or env.get("SUPABASE_URL", "")
SUPABASE_KEY      = env.get("SUPABASE_SERVICE_ROLE_KEY", "")

if not SUPABASE_URL or not SUPABASE_KEY:
    print("❌  NEXT_PUBLIC_SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY must be set in .env.local")
    sys.exit(1)

if not XLSX_PATH:
    print("Usage: python3 scripts/import-helyx-high.py <xlsx-path> [--run] [--campaign <name>]")
    sys.exit(1)

# ── Supabase helpers ──────────────────────────────────────────────────────────

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
}

def sb_get(table, params=""):
    url = f"{SUPABASE_URL}/rest/v1/{table}{params}"
    r = requests.get(url, headers=HEADERS)
    r.raise_for_status()
    return r.json()

def sb_post(table, payload, prefer="return=representation"):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    h = {**HEADERS, "Prefer": prefer}
    r = requests.post(url, headers=h, json=payload)
    r.raise_for_status()
    return r.json()

def sb_upsert(table, payload, on_conflict, prefer="resolution=merge-duplicates,return=representation"):
    url = f"{SUPABASE_URL}/rest/v1/{table}?on_conflict={on_conflict}"
    h = {**HEADERS, "Prefer": prefer}
    r = requests.post(url, headers=h, json=payload)
    r.raise_for_status()
    return r.json()

# ── Read Excel ────────────────────────────────────────────────────────────────

print(f"\n📂  Reading {XLSX_PATH} …")
wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

# Find sheet names (handles emoji)
high_sheet_name     = next(s for s in wb.sheetnames if "High" in s and "Emerging" in s)
contacts_sheet_name = next(s for s in wb.sheetnames if "Contact" in s)
print(f"   High sheet:     {high_sheet_name!r}")
print(f"   Contacts sheet: {contacts_sheet_name!r}")

# Roll Numbers from High sheet (row 1 = title, row 2 = headers, data from row 3)
ws_high = wb[high_sheet_name]
high_rolls = set()
for row in ws_high.iter_rows(min_row=3, values_only=True):
    if row[0] is not None:
        high_rolls.add(row[0])
print(f"\n🔵  High-priority Roll Numbers: {len(high_rolls)}")

# Contacts sheet (row 1 = title, row 2 = note, row 3 = headers, data from row 4)
# Columns (0-indexed): Roll#, CompanyName, CurrentPriority, PriorityChanged,
#   Seniority, FirstName, LastName, Title, Email, Phone, City, State, Country, LinkedIn
ws_contacts = wb[contacts_sheet_name]
contact_rows = []
for row in ws_contacts.iter_rows(min_row=4, values_only=True):
    if row[0] is None:
        continue
    roll = row[0]
    if roll not in high_rolls:
        continue
    contact_rows.append({
        "roll":       roll,
        "company":    (row[1] or "").strip(),
        "seniority":  (row[4] or "").strip(),
        "first":      (row[5] or "").strip(),
        "last":       (row[6] or "").strip(),
        "title":      (row[7] or "").strip(),
        "email":      (row[8] or "").strip(),
        "phone":      (row[9] or "").strip(),
        "city":       (row[10] or "").strip(),
        "state":      (row[11] or "").strip(),
        "country":    (row[12] or "").strip(),
        "linkedin":   (row[13] or "").strip(),
    })

print(f"👥  Contact rows matched:        {len(contact_rows)}")

# Build unique companies (Roll Number → company name + location from first contact)
companies = {}
for c in contact_rows:
    if c["roll"] not in companies:
        loc_parts = [p for p in [c["city"], c["state"], c["country"]] if p]
        companies[c["roll"]] = {
            "roll":     c["roll"],
            "name":     c["company"],
            "location": ", ".join(loc_parts) if loc_parts else None,
        }

print(f"🏢  Unique companies:             {len(companies)}")
print(f"📋  Contacts with email:          {sum(1 for c in contact_rows if c['email'])}")
print(f"🔗  Contacts with LinkedIn:       {sum(1 for c in contact_rows if c['linkedin'])}")

# Preview sample
print("\n── Sample companies ──────────────────────────────────────────────────────")
for co in list(companies.values())[:5]:
    print(f"   [{co['roll']}] {co['name']}  ({co['location'] or 'no location'})")

print("\n── Sample contacts ───────────────────────────────────────────────────────")
for c in contact_rows[:5]:
    name = f"{c['first']} {c['last']}".strip()
    print(f"   {name} | {c['title'][:40] if c['title'] else '—'} | {c['email'] or '—'}")

if DRY_RUN:
    print(f"""
╔══════════════════════════════════════════════════════════════╗
║  DRY RUN — nothing written                                   ║
║                                                              ║
║  Would insert:                                               ║
║    {len(companies):>5} organizations                                   ║
║    {len(contact_rows):>5} people                                         ║
║                                                              ║
║  Re-run with --run to execute, e.g.:                         ║
║    python3 scripts/import-helyx-high.py <xlsx> --run         ║
╚══════════════════════════════════════════════════════════════╝
""")
    sys.exit(0)

# ── Live import ───────────────────────────────────────────────────────────────

print(f"\n🚀  Importing into {SUPABASE_URL} …\n")

# 1. Fetch existing orgs by name to avoid duplicates (no unique constraint on name)
print("  Fetching existing organizations …")
existing_orgs_raw = sb_get("organizations", "?select=id,name&limit=10000")
existing_by_name = {r["name"].lower(): r["id"] for r in existing_orgs_raw}
print(f"  Found {len(existing_by_name)} existing organizations")

# 2. Upsert organizations
roll_to_org_id = {}
new_orgs = 0
for co in companies.values():
    name_key = co["name"].lower()
    if name_key in existing_by_name:
        roll_to_org_id[co["roll"]] = existing_by_name[name_key]
        continue
    result = sb_post("organizations", {
        "name":     co["name"],
        "location": co["location"],
        "source":   SOURCE_LABEL,
    })
    org_id = result[0]["id"] if isinstance(result, list) else result["id"]
    roll_to_org_id[co["roll"]] = org_id
    existing_by_name[name_key] = org_id
    new_orgs += 1

print(f"  ✅  Organizations — {new_orgs} new, {len(companies) - new_orgs} already existed")

# 3. Fetch existing people by linkedin_url for dedup
print("  Fetching existing people (LinkedIn URLs) …")
existing_people_raw = sb_get("people", "?select=id,linkedin_url,name,organization_id&limit=50000")
existing_by_linkedin = {
    r["linkedin_url"]: r["id"]
    for r in existing_people_raw
    if r.get("linkedin_url")
}
print(f"  Found {len(existing_by_linkedin)} existing people with LinkedIn URLs")

# 4. Upsert people
new_people = skipped = 0
for c in contact_rows:
    name = f"{c['first']} {c['last']}".strip()
    if not name:
        skipped += 1
        continue

    org_id   = roll_to_org_id.get(c["roll"])
    linkedin = c["linkedin"] or None

    # Normalize LinkedIn URL
    if linkedin:
        linkedin = re.sub(r"[?#].*$", "", linkedin).rstrip("/")
        if linkedin in existing_by_linkedin:
            skipped += 1
            continue

    payload = {
        "name":            name,
        "title":           c["title"] or None,
        "work_email":      c["email"] or None,
        "linkedin_url":    linkedin,
        "organization_id": org_id,
        "source":          SOURCE_LABEL,
    }

    try:
        result = sb_post("people", [payload])
        person_id = result[0]["id"] if result else None
    except Exception as e:
        # Unique constraint hit (same linkedin_url in source data, or re-run)
        skipped += 1
        continue

    if person_id:
        existing_by_linkedin[linkedin or ""] = person_id
        new_people += 1

print(f"  ✅  People — {new_people} inserted/updated, {skipped} skipped")

# 5. Campaign (optional)
if CAMPAIGN_ARG:
    print(f"\n  Setting up campaign: {CAMPAIGN_ARG!r} …")

    # Find or create campaign
    existing_campaigns = sb_get("campaigns", f"?name=eq.{requests.utils.quote(CAMPAIGN_ARG)}&select=id,name")
    if existing_campaigns:
        campaign_id = existing_campaigns[0]["id"]
        print(f"  ♻️   Reusing existing campaign {campaign_id}")
    else:
        result = sb_post("campaigns", {"name": CAMPAIGN_ARG})
        campaign_id = result[0]["id"] if isinstance(result, list) else result["id"]
        print(f"  ✅  Created campaign {campaign_id}")

    # Link organizations
    org_links = [
        {"campaign_id": campaign_id, "organization_id": oid}
        for oid in roll_to_org_id.values()
    ]
    # Batch in chunks of 500 to stay within Supabase payload limits
    CHUNK = 500
    org_linked = 0
    for i in range(0, len(org_links), CHUNK):
        sb_upsert("campaign_organizations", org_links[i:i+CHUNK],
                  on_conflict="campaign_id,organization_id")
        org_linked += len(org_links[i:i+CHUNK])
    print(f"  ✅  Linked {org_linked} organizations to campaign")

    # Link people — need to re-query people by org_id to get IDs
    people_in_orgs = sb_get(
        "people",
        f"?organization_id=in.({','.join(roll_to_org_id.values())})"
        f"&select=id&limit=10000"
    )
    people_links = [
        {"campaign_id": campaign_id, "person_id": p["id"]}
        for p in people_in_orgs
    ]
    people_linked = 0
    for i in range(0, len(people_links), CHUNK):
        sb_upsert("campaign_people", people_links[i:i+CHUNK],
                  on_conflict="campaign_id,person_id")
        people_linked += len(people_links[i:i+CHUNK])
    print(f"  ✅  Linked {people_linked} people to campaign")

print(f"""
╔══════════════════════════════════════════════════════════════╗
║  Import complete                                             ║
║    {new_orgs:>5} new organizations                                  ║
║    {new_people:>5} people inserted/updated                           ║
{"║    " + str(org_linked if CAMPAIGN_ARG else 0).rjust(5) + " orgs linked to campaign                           ║" if CAMPAIGN_ARG else ""}
╚══════════════════════════════════════════════════════════════╝
""")
